import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from io import BytesIO

# AJUSTE 8c: Dicionários de tradução de modelos
MODEL_OPTIONS = {
    0: "Binomial",
    1: "Black & Scholes",
    2: "Bjerkund-Stensland",
    3: "Último Pregão",
    4: "Média dos últimos Pregões",
    5: "Monte Carlo"
}
MODEL_VOLATILITY = {
    0: "N/A",
    1: "Desvio Padrão"
}

def gerar_excel_final(dados, capital_social):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    orgaos = ['Conselho de Administração', 'Diretoria Estatutária', 'Conselho Fiscal']

    def get_t(orgao):
        return ("*administração*" if "Administração" in orgao
                else ("*fiscal*" if "Fiscal" in orgao else "*diretoria*"))

    def get_col(i):
        return openpyxl.utils.get_column_letter(i)

    df_membros  = dados['membros']
    df_outorga  = dados['df_outorga']
    df_evid_85  = dados['8.5']
    df_evid_86  = dados['8.6']
    df_evid_87  = dados['8.7']
    df_evid_88  = dados['8.8']
    df_evid_89  = dados['8.9']
    df_evid_810 = dados['8.10']
    df_evid_811 = dados['8.11']

    # ==========================================
    # ABAS DE EVIDÊNCIAS
    # ==========================================
    col_org_membros = 'Orgão' if 'Orgão' in df_membros.columns else 'Órgão Administrativo'
    colunas_membros_export = [
        col_org_membros, 'NOME COMPLETO', 'DATA DE ENTRADA', 'DATA DE SAÍDA',
        'Pro_Rata_2025', 'Pro_Rata_2026',
        'Tem_85_2025', 'Tem_85_2026', 'Tem_86', 'Tem_87',
        'Tem_88', 'Tem_89', 'Tem_810', 'Tem_811'
    ]

    max_rows = {}
    for nome_aba, df_e in [
        ('Evid_Membros', df_membros[colunas_membros_export]),
        ('Evid_85',  df_evid_85),
        ('Evid_86',  df_evid_86),
        ('Evid_87',  df_evid_87),
        ('Evid_89',  df_evid_89),
        ('Evid_810', df_evid_810),
    ]:
        ws_e = wb.create_sheet(nome_aba)
        for r in dataframe_to_rows(df_e, index=False, header=True):
            ws_e.append(r)
        max_rows[nome_aba] = ws_e.max_row

    # Evid_88: A=Órgão B=Nome C=Programa D=Lote E=Data F=Qtd G=Preço_Ex H=Preço_Merc I=Ganho
    ws_e88 = wb.create_sheet('Evid_88')
    ws_e88.append(list(df_evid_88.columns) + ['Ganho_Bruto_Formula'])
    for i, r in enumerate(dataframe_to_rows(df_evid_88, index=False, header=False), start=2):
        ws_e88.append(list(r) + [f"=F{i}*(H{i}-G{i})"])
    max_rows['Evid_88'] = ws_e88.max_row

    # Evid_811: A=Órgão B=Nome C=Programa D=Lote E=Data F=Qtd G=Preço_Aq H=Preço_Merc I=Era_Estat J=Ganho
    ws_e811 = wb.create_sheet('Evid_811')
    ws_e811.append(list(df_evid_811.columns) + ['Ganho_Bruto_Formula'])
    for i, r in enumerate(dataframe_to_rows(df_evid_811, index=False, header=False), start=2):
        ws_e811.append(list(r) + [f"=F{i}*(H{i}-G{i})"])
    max_rows['Evid_811'] = ws_e811.max_row

    # ==========================================
    # HELPER: prazos com nome do programa (AJUSTE 4)
    # ==========================================
    def formatar_prazos(df_filtrado, tipo_data, acoes=False):
        if df_filtrado.empty:
            return "-"
        linhas = []
        for (dout, dalvo), grp in df_filtrado.groupby(['Data_Outorga', tipo_data]):
            if pd.isnull(dout) or pd.isnull(dalvo):
                continue
            nome_prog = grp['Programa'].iloc[0] if 'Programa' in grp.columns else dout.strftime('%d/%m/%Y')
            qtd = grp['Qtd' if acoes else 'Qtd_Saldo'].sum()
            linhas.append(
                f"{nome_prog}: {qtd:,.0f} {'ações' if acoes else 'opções'} até {dalvo.strftime('%d/%m/%Y')}".replace(',', '.')
            )
        return "\n".join(linhas) if linhas else "-"

    # ==========================================
    # QUADROS 8.5 e 8.9
    #
    # Evid_85/89 com Lote:
    #   A=Nome  B=Órgão  C=Programa  D=Lote  E=Preço  F=Qtd  G=Status
    # ==========================================
    def criar_quadro_mov(num, df_e, max_r, ano):
        ws = wb.create_sheet(f'Quadro_{num}')
        ws.append(['Grupos'] + orgaos)
        aba_evid = f'Evid_{num.replace("_2025","").replace("_2026","").replace(".","")}'
        labels = [
            "Nº total de membros", "Nº de membros remunerados",
            "Diluição potencial (%)", "Esclarecimento",
            "Em aberto no início do exercício (R$)",
            "Perdidas e expiradas no exercício (R$)",
            "Exercidas no exercício (R$)",
            "Em aberto no final do exercício (R$)",
        ]
        for i, lbl in enumerate(labels, start=2):
            ws.append([lbl])
            for c_idx, org in enumerate(orgaos):
                cel = f"{get_col(c_idx + 2)}{i}"
                mr_mem = max_rows['Evid_Membros']

                if "total de membros" in lbl:
                    col_p = "E" if ano == 2025 else "F"
                    ws[cel] = (f'=SUMIF(Evid_Membros!$A$2:$A${mr_mem}, "{get_t(org)}", '
                               f'Evid_Membros!${col_p}$2:${col_p}${mr_mem})')
                    ws[cel].number_format = '0.00'

                elif "remunerados" in lbl:
                    col_p    = "E" if ano == 2025 else "F"
                    col_flag = "G" if ano == 2025 else "H"
                    ws[cel] = (f'=SUMPRODUCT(--(Evid_Membros!$A$2:$A${mr_mem}="{get_t(org)}"), '
                               f'Evid_Membros!${col_p}$2:${col_p}${mr_mem}, '
                               f'--(Evid_Membros!${col_flag}$2:${col_flag}${mr_mem}=1))')
                    ws[cel].number_format = '0.00'

                elif "R$" in lbl:
                    st = ("Inicial " + str(ano) if "início" in lbl
                          else ("Final " + str(ano) if "final" in lbl
                                else ("Perdidas " + str(ano) if "Perdidas" in lbl
                                      else "Exercidas " + str(ano))))
                    # Evid_85/89: B=Órgão  E=Preço  F=Qtd  G=Status
                    ws[cel] = (f'=IFERROR('
                               f'SUMPRODUCT(--({aba_evid}!$B$2:$B${max_r}="{org}"), '
                               f'--({aba_evid}!$G$2:$G${max_r}="{st}"), '
                               f'{aba_evid}!$E$2:$E${max_r}, '
                               f'{aba_evid}!$F$2:$F${max_r}) / '
                               f'SUMIFS({aba_evid}!$F$2:$F${max_r}, '
                               f'{aba_evid}!$B$2:$B${max_r}, "{org}", '
                               f'{aba_evid}!$G$2:$G${max_r}, "{st}"), 0)')
                    ws[cel].number_format = 'R$ #,##0.00'

                elif "Diluição" in lbl:
                    # Evid_85/89: B=Órgão  F=Qtd  G=Status
                    ws[cel] = (f'=IFERROR('
                               f'SUMIFS({aba_evid}!$F$2:$F${max_r}, '
                               f'{aba_evid}!$B$2:$B${max_r}, "{org}", '
                               f'{aba_evid}!$G$2:$G${max_r}, "Final {ano}") / {capital_social}, 0)')
                    ws[cel].number_format = '0.0000%'

                else:
                    ws[cel] = "-"

    criar_quadro_mov('8.5_2025', df_evid_85, max_rows['Evid_85'], 2025)
    criar_quadro_mov('8.5_2026', df_evid_85, max_rows['Evid_85'], 2026)
    criar_quadro_mov('8.9_2025', df_evid_89, max_rows['Evid_89'], 2025)
    criar_quadro_mov('8.9_2026', df_evid_89, max_rows['Evid_89'], 2026)

    # ==========================================
    # QUADROS 8.6 e 8.10
    #
    # Evid_86 com Lote:
    #   A=Coluna_Rel  B=Órgão  C=Nome  D=Programa  E=Lote
    #   F=DataOut  G=DataCar  H=DataExp  I=Qtd_Outorgada  J=Fair_Value
    # Evid_810 com Lote:
    #   A=Coluna_Rel  B=Órgão  C=Nome  D=Programa  E=Lote
    #   F=DataOut  G=DataCar  H=Qtd_Outorgada  I=Fair_Value
    # ==========================================
    def criar_quadro_outorga(num, df_e, max_r):
        ws = wb.create_sheet(f'Quadro_{num}')
        cols_rel = ([c for c in df_e['Coluna_Relatorio'].unique() if pd.notnull(c)]
                    if not df_e.empty else [])
        ws.append(['Detalhamento das Outorgas'] + list(cols_rel))
        lbls = [
            "Nº total de membros", "N° de membros remunerados", "Data de outorga",
            "Quantidade outorgada", "Prazo para opções/ações", "Prazo máximo",
            "Prazo de restrição", "Valor justo na data", "Multiplicação",
        ]
        aba = f'Evid_{num.replace(".", "")}'
        col_flag_rem = "I" if num == '8.6' else "M"
        mr_mem = max_rows['Evid_Membros']

        for i, lbl in enumerate(lbls, start=2):
            ws.append([lbl])
            for c_idx, col_n in enumerate(cols_rel):
                cel = f"{get_col(c_idx + 2)}{i}"
                df_g = df_e[df_e['Coluna_Relatorio'] == col_n]
                if df_g.empty:
                    ws[cel] = "-"
                    continue

                org_val = df_g['Órgão Administrativo'].iloc[0]

                if "total de membros" in lbl:
                    ws[cel] = (f'=SUMIF(Evid_Membros!$A$2:$A${mr_mem}, "{get_t(org_val)}", '
                               f'Evid_Membros!$E$2:$E${mr_mem})')
                    ws[cel].number_format = '0.00'

                elif "remunerados" in lbl:
                    ws[cel] = (f'=SUMPRODUCT(--(Evid_Membros!$A$2:$A${mr_mem}="{get_t(org_val)}"), '
                               f'Evid_Membros!$E$2:$E${mr_mem}, '
                               f'--(Evid_Membros!${col_flag_rem}$2:${col_flag_rem}${mr_mem}=1))')
                    ws[cel].number_format = '0.00'

                elif "Data de outorga" in lbl:
                    ws[cel] = df_g['Data Outorga'].iloc[0].strftime('%d/%m/%Y') if pd.notnull(df_g['Data Outorga'].iloc[0]) else "-"

                elif "Quantidade" in lbl:
                    # 8.10: col I=Qtd_Outorgada  |  8.6: col I=Qtd_Outorgada
                    col_q = "$I" if num == '8.10' else "$I"
                    ws[cel] = f'=SUMIF({aba}!$A$2:$A${max_r}, "{col_n}", {aba}!{col_q}$2:{col_q}${max_r})'
                    ws[cel].number_format = '#,##0'

                elif "Prazo para" in lbl:
                    ws[cel] = df_g['Data Carência'].iloc[0].strftime('%d/%m/%Y') if pd.notnull(df_g['Data Carência'].iloc[0]) else "-"

                elif "Prazo máximo" in lbl:
                    ws[cel] = (df_g['Data Expiração'].iloc[0].strftime('%d/%m/%Y')
                               if '8.6' in num and 'Data Expiração' in df_g.columns and pd.notnull(df_g['Data Expiração'].iloc[0]) else "-")

                elif "restrição" in lbl:
                    ws[cel] = "N/A"

                elif "Valor justo" in lbl:
                    # 8.6:  I=Qtd_Outorgada  J=Fair_Value
                    # 8.10: I=Qtd_Outorgada  I=Fair_Value → H=Qtd H e I=FV
                    col_q = "$I"; col_v = "$J"
                    if num == '8.10':
                        col_q = "$H"; col_v = "$I"
                    ws[cel] = (f'=IFERROR(SUMPRODUCT(--({aba}!$A$2:$A${max_r}="{col_n}"), '
                               f'{aba}!{col_q}$2:{col_q}${max_r}, {aba}!{col_v}$2:{col_v}${max_r}) / '
                               f'SUMIF({aba}!$A$2:$A${max_r}, "{col_n}", {aba}!{col_q}$2:{col_q}${max_r}), 0)')
                    ws[cel].number_format = 'R$ #,##0.00'

                elif "Multiplicação" in lbl:
                    col_q = "$I"; col_v = "$J"
                    if num == '8.10':
                        col_q = "$H"; col_v = "$I"
                    ws[cel] = (f'=SUMPRODUCT(--({aba}!$A$2:$A${max_r}="{col_n}"), '
                               f'{aba}!{col_q}$2:{col_q}${max_r}, {aba}!{col_v}$2:{col_v}${max_r})')
                    ws[cel].number_format = 'R$ #,##0.00'

    criar_quadro_outorga('8.6',  df_evid_86, max_rows['Evid_86'])
    criar_quadro_outorga('8.10', df_evid_810, max_rows['Evid_810'])

    # ==========================================
    # QUADRO 8.7
    #
    # Evid_87 com Lote:
    #   A=Órgão  B=Nome  C=Programa  D=Lote  E=Status_Vesting
    #   F=Qtd_Saldo  G=Preço  H=Fair_Value
    # ==========================================
    ws_87 = wb.create_sheet('Quadro_8.7')
    ws_87.append(['2025'] + orgaos + ['Total'])
    lbls_87 = [
        "Nº total de membros", "N° de membros remunerados",
        "Opções não exercíveis", "Quantidade (Não exercíveis)",
        "Data em que se tornarão", "Prazo máximo (Não exercíveis)",
        "Restrição (Não)", "Preço médio (Não)", "Valor justo (Não)",
        "Opções exercíveis", "Quantidade (Exercíveis)",
        "Prazo máximo (Exercíveis)", "Restrição (Exercíveis)",
        "Preço médio (Exercíveis)", "Valor justo (Exercíveis)",
        "Valor justo do TOTAL",
    ]
    mr_mem = max_rows['Evid_Membros']
    mr_87  = max_rows['Evid_87']

    for i, lbl in enumerate(lbls_87, start=2):
        ws_87.append([lbl])
        ws_87[f"A{i}"].alignment = Alignment(wrapText=True, vertical='top')
        for c_idx, org in enumerate(orgaos + ['Total']):
            cel = f"{get_col(c_idx + 2)}{i}"
            df_t = df_evid_87 if org == 'Total' else df_evid_87[df_evid_87['Órgão Administrativo'] == org]
            st = "Não exercível" if "Não" in lbl else "Exercível"

            if "total de membros" in lbl:
                ws_87[cel] = (f'=SUM(Evid_Membros!$E$2:$E${mr_mem})' if org == 'Total'
                              else f'=SUMIF(Evid_Membros!$A$2:$A${mr_mem}, "{get_t(org)}", Evid_Membros!$E$2:$E${mr_mem})')
                ws_87[cel].number_format = '0.00'

            elif "remunerados" in lbl:
                # col J = Tem_87
                ws_87[cel] = (
                    f'=SUMPRODUCT(Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$J$2:$J${mr_mem}=1))' if org == 'Total'
                    else f'=SUMPRODUCT(--(Evid_Membros!$A$2:$A${mr_mem}="{get_t(org)}"), '
                         f'Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$J$2:$J${mr_mem}=1))')
                ws_87[cel].number_format = '0.00'

            elif "Quantidade" in lbl:
                # E=Status_Vesting  F=Qtd_Saldo
                ws_87[cel] = (f'=SUMIF(Evid_87!$E$2:$E${mr_87}, "{st}", Evid_87!$F$2:$F${mr_87})' if org == 'Total'
                              else f'=SUMIFS(Evid_87!$F$2:$F${mr_87}, Evid_87!$A$2:$A${mr_87}, "{org}", Evid_87!$E$2:$E${mr_87}, "{st}")')
                ws_87[cel].number_format = '#,##0'

            elif "Data" in lbl:
                ws_87[cel] = formatar_prazos(df_t[df_t['Status_Vesting'] == 'Não exercível'], 'Data_Carência')

            elif "Prazo" in lbl:
                ws_87[cel] = formatar_prazos(df_t[df_t['Status_Vesting'] == st], 'Data_Expiração')

            elif "Preço" in lbl or "Valor justo (" in lbl:
                # G=Preço  H=Fair_Value  F=Qtd_Saldo  E=Status_Vesting
                col_v = "$G" if "Preço" in lbl else "$H"
                n = (f'SUMPRODUCT(--(Evid_87!$E$2:$E${mr_87}="{st}"), Evid_87!$F$2:$F${mr_87}, Evid_87!{col_v}$2:{col_v}${mr_87})' if org == 'Total'
                     else f'SUMPRODUCT(--(Evid_87!$A$2:$A${mr_87}="{org}"), --(Evid_87!$E$2:$E${mr_87}="{st}"), Evid_87!$F$2:$F${mr_87}, Evid_87!{col_v}$2:{col_v}${mr_87})')
                d = (f'SUMIF(Evid_87!$E$2:$E${mr_87}, "{st}", Evid_87!$F$2:$F${mr_87})' if org == 'Total'
                     else f'SUMIFS(Evid_87!$F$2:$F${mr_87}, Evid_87!$A$2:$A${mr_87}, "{org}", Evid_87!$E$2:$E${mr_87}, "{st}")')
                ws_87[cel] = f'=IFERROR({n} / {d}, 0)'
                ws_87[cel].number_format = 'R$ #,##0.00'

            elif "TOTAL" in lbl:
                # F=Qtd_Saldo  H=Fair_Value
                ws_87[cel] = (f'=SUMPRODUCT(Evid_87!$F$2:$F${mr_87}, Evid_87!$H$2:$H${mr_87})' if org == 'Total'
                              else f'=SUMPRODUCT(--(Evid_87!$A$2:$A${mr_87}="{org}"), Evid_87!$F$2:$F${mr_87}, Evid_87!$H$2:$H${mr_87})')
                ws_87[cel].number_format = 'R$ #,##0.00'

            else:
                ws_87[cel] = "N/A" if "restrição" in lbl.lower() else ""

    # ==========================================
    # QUADRO 8.8
    #
    # Evid_88 com Lote:
    #   A=Órgão  B=Nome  C=Programa  D=Lote  E=Data  F=Qtd  G=Preço_Ex  H=Preço_Merc  I=Ganho
    # ==========================================
    ws_88 = wb.create_sheet('Quadro_8.8')
    ws_88.append(['2025'] + orgaos + ['Total'])
    lbls_88 = ["Nº total de membros", "N° de membros remunerados", "Número de ações",
               "Preço médio de exercício", "Preço médio de mercado", "Multiplicação (Ganho Total)"]
    mr_88 = max_rows['Evid_88']

    for i, lbl in enumerate(lbls_88, start=2):
        ws_88.append([lbl])
        for c_idx, org in enumerate(orgaos + ['Total']):
            cel = f"{get_col(c_idx + 2)}{i}"

            if "total de membros" in lbl:
                ws_88[cel] = (f'=SUM(Evid_Membros!$E$2:$E${mr_mem})' if org == 'Total'
                              else f'=SUMIF(Evid_Membros!$A$2:$A${mr_mem}, "{get_t(org)}", Evid_Membros!$E$2:$E${mr_mem})')
                ws_88[cel].number_format = '0.00'

            elif "remunerados" in lbl:
                # col K = Tem_88
                ws_88[cel] = (
                    f'=SUMPRODUCT(Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$K$2:$K${mr_mem}=1))' if org == 'Total'
                    else f'=SUMPRODUCT(--(Evid_Membros!$A$2:$A${mr_mem}="{get_t(org)}"), '
                         f'Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$K$2:$K${mr_mem}=1))')
                ws_88[cel].number_format = '0.00'

            elif "Número" in lbl:
                # F=Qtd
                ws_88[cel] = (f'=SUM(Evid_88!$F$2:$F${mr_88})' if org == 'Total'
                              else f'=SUMIF(Evid_88!$A$2:$A${mr_88}, "{org}", Evid_88!$F$2:$F${mr_88})')
                ws_88[cel].number_format = '#,##0'

            elif "exercício" in lbl or "mercado" in lbl:
                # G=Preço_Ex  H=Preço_Merc  F=Qtd
                col_v = "$G" if "exercício" in lbl else "$H"
                n = (f'SUMPRODUCT(Evid_88!$F$2:$F${mr_88}, Evid_88!{col_v}$2:{col_v}${mr_88})' if org == 'Total'
                     else f'SUMPRODUCT(--(Evid_88!$A$2:$A${mr_88}="{org}"), Evid_88!$F$2:$F${mr_88}, Evid_88!{col_v}$2:{col_v}${mr_88})')
                d = (f'SUM(Evid_88!$F$2:$F${mr_88})' if org == 'Total'
                     else f'SUMIF(Evid_88!$A$2:$A${mr_88}, "{org}", Evid_88!$F$2:$F${mr_88})')
                ws_88[cel] = f'=IFERROR({n} / {d}, 0)'
                ws_88[cel].number_format = 'R$ #,##0.00'

            elif "Multiplicação" in lbl:
                # I=Ganho
                ws_88[cel] = (f'=SUM(Evid_88!$I$2:$I${mr_88})' if org == 'Total'
                              else f'=SUMIF(Evid_88!$A$2:$A${mr_88}, "{org}", Evid_88!$I$2:$I${mr_88})')
                ws_88[cel].number_format = 'R$ #,##0.00'

    # ==========================================
    # QUADRO 8.11
    #
    # Evid_811 com Lote:
    #   A=Órgão  B=Nome  C=Programa  D=Lote  E=Data  F=Qtd
    #   G=Preço_Aq  H=Preço_Merc  I=Era_Estat  J=Ganho
    # ==========================================
    ws_811 = wb.create_sheet('Quadro_8.11')
    ws_811.append(['2025'] + orgaos + ['Total'])
    lbls_811 = [
        "Nº total de membros", "N° de membros remunerados", "Número de ações",
        "Preço médio ponderado de aquisição",
        "Preço médio ponderado de mercado das ações adquiridas",
        "Multiplicação do total das ações adquiridas",
        "N° de membros que eram Estatutários na data do exercício",
    ]
    mr_811 = max_rows['Evid_811']

    for i, lbl in enumerate(lbls_811, start=2):
        ws_811.append([lbl])
        ws_811[f"A{i}"].alignment = Alignment(wrapText=True, vertical='top')
        for c_idx, org in enumerate(orgaos + ['Total']):
            cel = f"{get_col(c_idx + 2)}{i}"

            if "total de membros" in lbl:
                ws_811[cel] = (f'=SUM(Evid_Membros!$E$2:$E${mr_mem})' if org == 'Total'
                               else f'=SUMIF(Evid_Membros!$A$2:$A${mr_mem}, "{get_t(org)}", Evid_Membros!$E$2:$E${mr_mem})')
                ws_811[cel].number_format = '0.00'

            elif "remunerados" in lbl:
                # col N = Tem_811
                ws_811[cel] = (
                    f'=SUMPRODUCT(Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$N$2:$N${mr_mem}=1))' if org == 'Total'
                    else f'=SUMPRODUCT(--(Evid_Membros!$A$2:$A${mr_mem}="{get_t(org)}"), '
                         f'Evid_Membros!$E$2:$E${mr_mem}, --(Evid_Membros!$N$2:$N${mr_mem}=1))')
                ws_811[cel].number_format = '0.00'

            elif "Número de ações" in lbl:
                # F=Qtd
                ws_811[cel] = (f'=SUM(Evid_811!$F$2:$F${mr_811})' if org == 'Total'
                               else f'=SUMIF(Evid_811!$A$2:$A${mr_811}, "{org}", Evid_811!$F$2:$F${mr_811})')
                ws_811[cel].number_format = '#,##0'

            elif "Multiplicação" in lbl:
                # J=Ganho
                ws_811[cel] = (f'=SUM(Evid_811!$J$2:$J${mr_811})' if org == 'Total'
                               else f'=SUMIF(Evid_811!$A$2:$A${mr_811}, "{org}", Evid_811!$J$2:$J${mr_811})')
                ws_811[cel].number_format = 'R$ #,##0.00'

            elif "aquisição" in lbl or "mercado" in lbl:
                # G=Preço_Aq  H=Preço_Merc  F=Qtd
                col_v = "$G" if "aquisição" in lbl else "$H"
                n = (f'SUMPRODUCT(Evid_811!$F$2:$F${mr_811}, Evid_811!{col_v}$2:{col_v}${mr_811})' if org == 'Total'
                     else f'SUMPRODUCT(--(Evid_811!$A$2:$A${mr_811}="{org}"), Evid_811!$F$2:$F${mr_811}, Evid_811!{col_v}$2:{col_v}${mr_811})')
                d = (f'SUM(Evid_811!$F$2:$F${mr_811})' if org == 'Total'
                     else f'SUMIF(Evid_811!$A$2:$A${mr_811}, "{org}", Evid_811!$F$2:$F${mr_811})')
                ws_811[cel] = f'=IFERROR({n} / {d}, 0)'
                ws_811[cel].number_format = 'R$ #,##0.00'

            elif "Estatutários" in lbl:
                # I=Era_Estatutario
                ws_811[cel] = (f'=COUNTIF(Evid_811!$I$2:$I${mr_811}, TRUE)' if org == 'Total'
                               else f'=COUNTIFS(Evid_811!$A$2:$A${mr_811}, "{org}", Evid_811!$I$2:$I${mr_811}, TRUE)')
                ws_811[cel].number_format = '0'

    # ==========================================
    # QUADRO 8.12
    # AJUSTE 8a: outorgas em ABERTO no início de 2025 (Data de Outorga < 2025)
    # ==========================================
    ws_812 = wb.create_sheet('Quadro_8.12')
    df_outorga['Data de Outorga'] = pd.to_datetime(df_outorga['Data de Outorga'], errors='coerce', dayfirst=True)
    df_25 = df_outorga[df_outorga['Data de Outorga'].dt.year < 2025].copy()

    if df_25.empty:
        ws_812.append(["Variável", "Nenhum programa em aberto no início de 2025"])
    else:
        df_25['Chave_Coluna'] = df_25['Programa'] + " (" + df_25['Órgão Administrativo'] + ")"
        col_progs = df_25['Chave_Coluna'].unique()
        ws_812.append(["Variável"] + list(col_progs))

        labels_812 = [
            "Modelo de Precificação",
            "Preço Médio Ponderado das Ações (R$)",
            "Preço de Exercício (R$)",
            "Volatilidade Esperada (%)",
            "Prazo de vida da opção",
            "Dividendos Esperados (%)",
            "Taxa de juros livre de riscos (%)",
            "Método utilizado (exercício antecipado)",
            "Forma de determinação da volatilidade",
            "Outra característica incorporada",
        ]
        for i, lbl in enumerate(labels_812, start=2):
            linha = [lbl]
            ws_812.column_dimensions['A'].width = 50
            ws_812[f"A{i}"].alignment = Alignment(wrapText=True, vertical='top')

            for col_n in col_progs:
                dp = df_25[df_25['Chave_Coluna'] == col_n].iloc[0]

                if lbl == "Modelo de Precificação":
                    modelo_raw = pd.to_numeric(str(dp.get('Model Options', '')).replace(',', '.'), errors='coerce')
                    val = MODEL_OPTIONS.get(int(modelo_raw), "A preencher") if pd.notnull(modelo_raw) else "A preencher"

                elif "Preço Médio Ponderado" in lbl:
                    v = pd.to_numeric(str(dp.get('Preço da Ação / Opção', 0)).replace(',', '.'), errors='coerce')
                    val = f"R$ {v:,.2f}".replace('.', ',') if pd.notnull(v) and v != 0 else "A preencher"

                elif "Preço de Exercício" in lbl:
                    v = pd.to_numeric(str(dp.get('Preço de Exercício na Outorga', 0)).replace(',', '.'), errors='coerce')
                    val = f"R$ {v:,.2f}".replace('.', ',') if pd.notnull(v) and v != 0 else "-"

                elif "Volatilidade Esperada" in lbl:
                    v = pd.to_numeric(str(dp.get('Volatilidade', 0)).replace(',', '.'), errors='coerce')
                    val = f"{v*100:,.2f}%".replace('.', ',') if pd.notnull(v) and v != 0 else "A preencher"

                elif "Prazo de vida" in lbl:
                    dout = dp.get('Data de Outorga')
                    dexp = pd.to_datetime(dp.get('Data de Expiração'), errors='coerce', dayfirst=True)
                    if pd.notnull(dout) and pd.notnull(dexp):
                        if dexp.year >= 9999:
                            dcar = pd.to_datetime(dp.get('Data da Carência'), errors='coerce', dayfirst=True)
                            val = f"{(dcar - dout).days / 365.25:.1f} anos" if pd.notnull(dcar) else "A preencher"
                        else:
                            val = f"{(dexp - dout).days / 365.25:.1f} anos"
                    else:
                        val = "A preencher"

                elif "Dividendos Esperados" in lbl:
                    v = pd.to_numeric(str(dp.get('Dividendos Esperados', 0)).replace(',', '.'), errors='coerce')
                    val = f"{v*100:,.2f}%".replace('.', ',') if pd.notnull(v) and v != 0 else "A preencher"

                elif "Taxa de juros" in lbl:
                    v = pd.to_numeric(str(dp.get('Taxa de Juros Livre de Risco', 0)).replace(',', '.'), errors='coerce')
                    val = f"{v*100:,.2f}%".replace('.', ',') if pd.notnull(v) and v != 0 else "A preencher"

                elif "antecipado" in lbl:
                    v = dp.get('Proporção de Exercício Antecipado', '')
                    val = str(v) if pd.notnull(v) and str(v).strip() not in ('', 'nan') else "A preencher"

                elif "volatilidade" in lbl.lower():
                    vol_raw = pd.to_numeric(str(dp.get('ModelVolatility', '')).replace(',', '.'), errors='coerce')
                    val = MODEL_VOLATILITY.get(int(vol_raw), "A preencher") if pd.notnull(vol_raw) else "A preencher"

                else:
                    val = "A preencher"

                linha.append(val)
            ws_812.append(linha)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
